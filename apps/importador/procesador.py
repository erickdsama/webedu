from openpyxl import load_workbook
from django.template.defaultfilters import title

from apps.escuela.forms import InstitucionForm
from apps.escuela.models import Alumno, InformacionPersonal, Grupo


class ProcesadorAlumnos(object):
    alumnos = list()
    alumnos_creados = 0
    alumnos_actualizados = 0

    def __init__(self, archivo):
        self.archivo = archivo

    def leerArchivo(self):
        wb = load_workbook(self.archivo)
        ws = wb.get_active_sheet()
        alumnos = list()
        encabezados = ("campus", "nivel", "grupo", "matricula", "apellido_paterno", "apellido_materno", "nombre")
        num_row = 0
        for row in ws.rows:
            if num_row != 0:
                i = 0
                alumno = dict()
                for encabezado, cell in zip(encabezados, row):
                    if cell.value is not None:
                        alumno.update({encabezado: cell.value})
                    i += 1
                alumnos.append(alumno)
            num_row += 1
        return alumnos

    def get_alumnos_list(self):
        if len(self.alumnos) == 0:
            self.alumnos = self.leerArchivo()
        return self.alumnos

    def guardar_alumnos(self):
        for alumno in self.get_alumnos_list():
            alumnoObj = Alumno.objects.filter(matricula=alumno.get('matricula'))
            if not alumnoObj.exists():
                # Crear alumno
                alumnoObj = Alumno.objects.create(
                    matricula=alumno.get('matricula'),
                    apellido_paterno=alumno.get('apellido_paterno').capitalize(),
                    apellido_materno=alumno.get('apellido_materno').capitalize(),
                    nombre=title(alumno.get('nombre')),
                    email="al{}@alumnos.uacj.mx".format(alumno.get('matricula')),
                    nivel=alumno.get('nivel'),
                    activo=True
                )
                InformacionPersonal.objects.create(alumno=alumnoObj)
                self.alumnos_creados += 1
            else:
                alumnoObj.update(nivel=alumno.get('nivel'))
                alumnoObj = alumnoObj.last()
                self.alumnos_actualizados += 1
            # Guardar en grupo
            grupo = Grupo.objects.filter(periodo__activo=True, grupo__iexact=alumno.get('grupo'),
                                         campus__iexact=alumno.get('campus'))

            if grupo.exists():
                grupo = grupo.last()
                grupo.alumnos.add(alumnoObj)


class ProcesadorAlumnos2(object):
    alumnos = list()
    alumnos_creados = 0
    alumnos_actualizados = 0

    def __init__(self, archivo):
        self.archivo = archivo

    def leerArchivo(self):
        wb = load_workbook(self.archivo)
        ws = wb.get_active_sheet()
        alumnos = list()
        # encabezados = ("matricu","nivel","grupo","matricula","apellido_paterno","apellido_materno","nombre")
        encabezados = ('matricula', 'email', 'nombre', 'nivel', 'telefono', 'telefono2', 'no_servicio_medico', 'tipo_seguro')
        num_row = 0
        for row in ws.rows:
            if num_row != 0:
                i = 0
                alumno = dict()
                for encabezado, cell in zip(encabezados, row):
                    if cell.value is not None:
                        alumno.update({encabezado: str(cell.value)})
                    i += 1
                alumnos.append(alumno)
            num_row += 1
        return alumnos

    def get_alumnos_list(self):
        if len(self.alumnos) == 0:
            self.alumnos = self.leerArchivo()
        return self.alumnos

    def guardar_alumnos(self):
        for alumno in self.get_alumnos_list():
            alumnoObj = Alumno.objects.filter(matricula=alumno.get('matricula'))
            if not alumnoObj.exists():
                # Crear alumno
                alumnoObj = Alumno.objects.create(
                    matricula=alumno.get('matricula'),
                    nombre=alumno.get('nombre').upper(),
                    email="al{}@alumnos.uacj.mx".format(alumno.get('matricula')),
                    nivel=alumno.get('nivel'),
                    activo=True
                )
                if alumno.get('telefono') and len(alumno.get('telefono')) > 10:
                    alumno.update({'telefono': alumno.get('telefono')[0:10]})
                if alumno.get('telefono2') and len(alumno.get('telefono2')) > 10:
                    alumno.update({'telefono2': alumno.get('telefono2')[0:10]})
                InformacionPersonal.objects.create(alumno=alumnoObj, telefono=alumno.get('telefono'),
                                                   telefono2=alumno.get('telefono2'),
                                                   no_servicio_medico=alumno.get('no_servicio_medico'),
                                                   tipo_seguro=alumno.get('tipo_seguro'))
                self.alumnos_creados += 1
            else:
                alumnoObj.update(nivel=alumno.get('nivel'))
                self.alumnos_actualizados += 1
            # Guardar en grupo
            # grupo = Grupo.objects.filter(periodo__activo=True, grupo__iexact=alumno.get('grupo'),campus__iexact=alumno.get('campus'))
            #
            # if grupo.exists():
            #     grupo = grupo.last()
            #     grupo.alumnos.add(alumnoObj)


class ReaderFile(object):
    """
    Clase utilizada para leer archivos de excel y crear una lista de diccionarios
    """
    encabezados = None
    data_file = []

    def __init__(self, archivo):
        self.archivo = archivo
        self.data_file = self.leerArchivo()

    def leerArchivo(self):
        wb = load_workbook(self.archivo)
        ws = wb.get_active_sheet()
        data_file = list()
        num_row = 0
        for row in ws.rows:
            if num_row != 0:
                i = 0
                obj = dict()
                for encabezado, cell in zip(self.encabezados, row):
                    if cell.value is not None:
                        obj.update({encabezado: str(cell.value)})
                    i += 1
                obj.update({'num_fila': num_row})
                data_file.append(obj)
            num_row += 1
        return data_file

    def get_data(self):
        return self.data_file


class ProcesadorInstituciones(ReaderFile):
    instituciones = list()
    instituciones_creadas = 0
    instituciones_actualizadas = 0
    errores = list()
    encabezados = ['nombre', 'direccion', 'referencia', 'telefono', 'contacto_directo', 'contacto_directo_email',
                   'responsable_principal', 'nivel', 'subnivel', 'ambito', 'sector']
    form_class = InstitucionForm

    def verificar_informacion(self):
        for row in self.data_file:
            form = self.form_class(data=row)
            if not form.is_valid():
                self.errores.append({'Fila {}'.format(row.get('num_fila')): form.errors})
        if len(self.errores) > 0:
            print("Errores encontrados")
            print(self.errores)

    def guardar_alumnos(self):

        for alumno in self.get_alumnos_list():
            alumnoObj = Alumno.objects.filter(matricula=alumno.get('matricula'))
            if not alumnoObj.exists():
                # Crear alumno
                alumnoObj = Alumno.objects.create(
                    matricula=alumno.get('matricula'),
                    nombre=alumno.get('nombre').upper(),
                    email="al{}@alumnos.uacj.mx".format(alumno.get('matricula')),
                    nivel=alumno.get('nivel'),
                    activo=True
                )
                if alumno.get('telefono') and len(alumno.get('telefono')) > 10:
                    alumno.update({'telefono': alumno.get('telefono')[0:10]})
                if alumno.get('telefono2') and len(alumno.get('telefono2')) > 10:
                    alumno.update({'telefono2': alumno.get('telefono2')[0:10]})
                InformacionPersonal.objects.create(alumno=alumnoObj, telefono=alumno.get('telefono'),
                                                   telefono2=alumno.get('telefono2'),
                                                   no_servicio_medico=alumno.get('no_servicio_medico'),
                                                   tipo_seguro=alumno.get('tipo_seguro'))
                self.alumnos_creados += 1
            else:
                alumnoObj.update(nivel=alumno.get('nivel'))
                self.alumnos_actualizados += 1
